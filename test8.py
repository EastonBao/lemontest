import requests
import openpyxl
import jsonpath
import pprint
#从表格中取对应的数据
def read_data(failname,sheetname):
    wb = openpyxl.load_workbook(failname)
    sheet = wb[sheetname]
    max_row = sheet.max_row
    case_list = []
    for i in range(2,max_row+1):
        case = dict(
        id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,
        head = sheet.cell(row=i,column=6).value,
        data = sheet.cell(row=i,column=7).value,
        expecte = sheet.cell(row=i,column=8).value
        )
        case_list.append(case)
    return case_list
#repason = read_data("test_test.xlsx", "register")

#利用取出的数据进行自动测试
def api_request(api_url,api_data,api_head):
    real_rult = requests.post(url=api_url, json=api_data, headers=api_head)
    return real_rult.json()
def expected(failname,sheetname):
    repason = read_data(failname, sheetname)
    rult = []
    for cases in repason:
        api_id = cases.get("id")
        api_url = cases.get("url")
        api_data = eval(cases.get("data"))
        api_head = eval(cases.get("head"))
        api_expecte = eval(cases.get("expecte"))
        #print(type(api_expecte))
        #print(api_data,api_url,api_expecte,api_id)
        #print(cases)
        real_rult1 = api_request(api_url, api_data, api_head)
        #print(type(real_rult))
        #pprint.pprint(real_rult)
        res_expecte = api_expecte["msg"]
        res_rult = real_rult1["msg"]
        # print(type(res_rult))
        if res_rult == res_expecte:
            rult_r = "通过"
            print("第{}个测试用例执行通过".format(api_id))
        else:
            rult_r = "不通过"
            print(("第{}个测试用例执行不通过".format(api_id)))
        rult.append(rult_r)
    return rult
def write(failname,sheetname,row,column):
    rult = expected(failname,sheetname)
    wb =openpyxl.load_workbook(failname)
    sheet = wb[sheetname]
    a = row
    for i in rult:
        #print(i)
        sheet.cell(row=a, column=column).value = i
        wb.save(failname)
        a += 1
write("test_test.xlsx", "register", 2, 9)

